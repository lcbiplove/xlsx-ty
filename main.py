import openpyxl, glob
 
def update(final_sheet_obj, row, date, tax_num, po_num, total):
    sh_date = final_sheet_obj.cell(row, 1)
    sh_date.value = date
    sh_num = final_sheet_obj.cell(row, 2)
    sh_num.value = tax_num
    sh_po_num = final_sheet_obj.cell(row, 4)
    sh_po_num.value = po_num
    sh_total = final_sheet_obj.cell(row, 6)
    sh_total.value = total

def main():
    starting_row = 15

    starting_num = int(input("ENTER STARTING NUM: "))
    ending_num = int(input("ENTER ENDING NUM: "))
    filename = input("Enter filename including format (.xlsx): ")

    statement_block = openpyxl.load_workbook(filename)
    final_sheet_obj = statement_block.active

    for path in glob.glob('* - TYROOLA *.xlsx'):
        try:
            ty_num = int(path.split(" ")[0])

            if ty_num < starting_num or ty_num > ending_num:
                continue

            print("STARTING FILE: ", path)
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            
            po_num = sheet_obj.cell(row = 4, column = 5).value
            date = sheet_obj.cell(row = 3, column = 3).value
            total = sheet_obj.cell(row = 35, column = 6).value

            tax_value = sheet_obj.cell(row = 2, column = 1).value
            tax_num = tax_value.split("#")[1].strip()

            update(final_sheet_obj, starting_row, date, tax_num, po_num, total)
            starting_row += 1
            print("ENDED FILE: ", path)
        except:
            print("Could not find num")
        
    statement_block.save(filename)
    wb_obj.close()
    statement_block.close()
    print("DONE")


if __name__ == "__main__":
    main()
